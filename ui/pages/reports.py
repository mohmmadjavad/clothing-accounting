from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QFrame,
    QComboBox, QMessageBox, QFileDialog, QTabWidget,
    QSizePolicy, QScrollArea, QAbstractItemView
)
from PyQt6.QtCore import Qt, QSize
from PyQt6.QtGui import QFont, QColor, QBrush, QPainter, QPen, QLinearGradient
from ui.styles import ModernStyle
from utils.jdatetime_utils import get_current_shamsi_date
import jdatetime


# ─────────────────────────────────────────────
#  نمودار میله‌ای سفارشی با QPainter
# ─────────────────────────────────────────────
class BarChartWidget(QWidget):
    MONTHS = ['فروردین','اردیبهشت','خرداد','تیر','مرداد','شهریور',
              'مهر','آبان','آذر','دی','بهمن','اسفند']
    COLORS = ['#6C5CE7','#00CEC9','#FD79A8','#00B894',
              '#FDCB6E','#74B9FF','#E17055','#A29BFE',
              '#55EFC4','#0984E3','#E84393','#636E72']

    def __init__(self, parent=None):
        super().__init__(parent)
        self.data = []          # list of {'month','amount','count'}
        self.setMinimumHeight(220)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.setStyleSheet("background: white; border-radius: 12px;")

    def set_data(self, data):
        self.data = data
        self.update()

    def paintEvent(self, event):
        if not self.data:
            return
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        W, H = self.width(), self.height()
        pad_l, pad_r, pad_t, pad_b = 20, 20, 20, 50

        max_val = max(d['amount'] for d in self.data) or 1
        n = len(self.data)
        slot_w = (W - pad_l - pad_r) / n
        bar_w = max(6, slot_w * 0.55)

        for i, d in enumerate(self.data):
            x = pad_l + i * slot_w + (slot_w - bar_w) / 2
            bar_h = max(4, (d['amount'] / max_val) * (H - pad_t - pad_b))
            y = H - pad_b - bar_h

            # گرادیان میله
            grad = QLinearGradient(x, y, x, y + bar_h)
            color = QColor(self.COLORS[i % len(self.COLORS)])
            lighter = color.lighter(130)
            grad.setColorAt(0, lighter)
            grad.setColorAt(1, color)

            painter.setBrush(grad)
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawRoundedRect(int(x), int(y), int(bar_w), int(bar_h), 4, 4)

            # مبلغ بالای میله
            if d['amount'] > 0:
                painter.setPen(QPen(QColor(ModernStyle.TEXT_SECONDARY)))
                painter.setFont(QFont("Tahoma", 7))
                amt_text = f"{int(d['amount']/1000):,}K" if d['amount'] >= 1000 else str(int(d['amount']))
                painter.drawText(int(x), int(y) - 3, int(bar_w), 14,
                                 Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignBottom, amt_text)

            # نام ماه
            painter.setPen(QPen(QColor(ModernStyle.TEXT_PRIMARY)))
            painter.setFont(QFont("Tahoma", 8))
            painter.drawText(int(x) - 4, H - pad_b + 6, int(bar_w) + 8, 20,
                             Qt.AlignmentFlag.AlignHCenter, d['month'][:3])

        painter.end()


# ─────────────────────────────────────────────
#  کارت خلاصه آمار
# ─────────────────────────────────────────────
class SummaryCard(QFrame):
    def __init__(self, icon, label, value, color, parent=None):
        super().__init__(parent)
        self.setObjectName("card")
        self.setStyleSheet(f"""
            QFrame#card {{
                background: white;
                border-radius: 14px;
                border: 1px solid {ModernStyle.BORDER};
            }}
        """)
        self.setMinimumHeight(90)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 12, 16, 12)
        lay.setSpacing(4)

        top = QHBoxLayout()
        ico = QLabel(icon)
        ico.setFont(QFont("Segoe UI Emoji", 20))
        ico.setStyleSheet("background:transparent;border:none;")
        top.addWidget(ico)
        top.addStretch()

        self.val_lbl = QLabel(value)
        self.val_lbl.setFont(QFont("Tahoma", 16, QFont.Weight.Bold))
        self.val_lbl.setStyleSheet(f"color:{color};background:transparent;border:none;")
        self.val_lbl.setAlignment(Qt.AlignmentFlag.AlignRight)
        top.addWidget(self.val_lbl)
        lay.addLayout(top)

        lbl = QLabel(label)
        lbl.setFont(QFont("Tahoma", 10))
        lbl.setStyleSheet(f"color:{ModernStyle.TEXT_SECONDARY};background:transparent;border:none;")
        lay.addWidget(lbl)

    def set_value(self, v):
        self.val_lbl.setText(str(v))


# ─────────────────────────────────────────────
#  صفحه اصلی گزارشات
# ─────────────────────────────────────────────
class ReportsPage(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self._init_ui()
        self.load_all()

    # ── ساخت UI ──────────────────────────────
    def _init_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── نوار هدر بنفش ──
        header_frame = QFrame()
        header_frame.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 {ModernStyle.PRIMARY}, stop:1 #8e7cf0);
                border-radius: 0px;
            }}
        """)
        header_frame.setFixedHeight(72)
        hlay = QHBoxLayout(header_frame)
        hlay.setContentsMargins(28, 0, 28, 0)

        title_lbl = QLabel("📈  گزارش‌گیری و آمار")
        title_lbl.setFont(QFont("Tahoma", 18, QFont.Weight.Bold))
        title_lbl.setStyleSheet("color:white;background:transparent;border:none;")
        hlay.addWidget(title_lbl)
        hlay.addStretch()

        hlay.addWidget(self._lbl("سال:", "white", 11))
        self.year_combo = QComboBox()
        self.year_combo.setFixedWidth(90)
        self.year_combo.setFixedHeight(36)
        self.year_combo.setStyleSheet("""
            QComboBox {
                background: rgba(255,255,255,0.22);
                color: white;
                border: 1px solid rgba(255,255,255,0.4);
                border-radius: 8px;
                padding: 4px 10px;
                font-size: 13px;
            }
            QComboBox::drop-down { border:none; }
            QComboBox QAbstractItemView {
                background: white; color: #2D3436;
                selection-background-color: #6C5CE7;
                selection-color: white;
            }
        """)
        cy = jdatetime.datetime.now().year
        for y in range(cy - 2, cy + 1):
            self.year_combo.addItem(str(y), y)
        self.year_combo.setCurrentIndex(self.year_combo.count() - 1)
        self.year_combo.currentIndexChanged.connect(self.load_all)
        hlay.addWidget(self.year_combo)

        refresh_btn = QPushButton("🔄 به‌روزرسانی")
        refresh_btn.setFixedHeight(36)
        refresh_btn.setFixedWidth(130)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background: rgba(255,255,255,0.18);
                color: white;
                border: 1px solid rgba(255,255,255,0.4);
                border-radius: 8px;
                font-size: 12px;
                padding: 0 12px;
            }
            QPushButton:hover { background: rgba(255,255,255,0.30); }
        """)
        refresh_btn.clicked.connect(self.load_all)
        hlay.addWidget(refresh_btn)
        root.addWidget(header_frame)

        # ── محتوای اصلی ──
        body_scroll = QScrollArea()
        body_scroll.setWidgetResizable(True)
        body_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        body_scroll.setStyleSheet("QScrollArea{border:none;background:#F5F6FA;}")
        root.addWidget(body_scroll, stretch=1)

        body = QWidget()
        body.setStyleSheet("background:#F5F6FA;")
        body_scroll.setWidget(body)

        self.body_layout = QVBoxLayout(body)
        self.body_layout.setContentsMargins(24, 24, 24, 24)
        self.body_layout.setSpacing(20)

        # ── کارت‌های خلاصه (ردیف بالا) ──
        cards_row = QHBoxLayout()
        cards_row.setSpacing(14)
        self.card_total_orders   = SummaryCard("📋", "کل سفارشات سال",   "—", ModernStyle.PRIMARY)
        self.card_total_revenue  = SummaryCard("💰", "کل درآمد سال",      "—", ModernStyle.SUCCESS)
        self.card_avg_order      = SummaryCard("📊", "میانگین هر سفارش",  "—", ModernStyle.SECONDARY)
        self.card_top_month      = SummaryCard("🏆", "پربازده‌ترین ماه",  "—", ModernStyle.ACCENT)
        for c in [self.card_total_orders, self.card_total_revenue,
                  self.card_avg_order, self.card_top_month]:
            cards_row.addWidget(c)
        self.body_layout.addLayout(cards_row)

        # ── نمودار ──
        chart_frame = QFrame()
        chart_frame.setObjectName("card")
        chart_frame.setStyleSheet(f"""
            QFrame#card {{
                background: white;
                border-radius: 16px;
                border: 1px solid {ModernStyle.BORDER};
            }}
        """)
        chart_fl = QVBoxLayout(chart_frame)
        chart_fl.setContentsMargins(20, 16, 20, 16)
        chart_fl.setSpacing(10)

        chart_title_row = QHBoxLayout()
        ct = QLabel("📊  نمودار فروش ماهانه")
        ct.setFont(QFont("Tahoma", 13, QFont.Weight.Bold))
        ct.setStyleSheet(f"color:{ModernStyle.TEXT_PRIMARY};background:transparent;border:none;")
        chart_title_row.addWidget(ct)
        chart_title_row.addStretch()
        chart_fl.addLayout(chart_title_row)

        self.bar_chart = BarChartWidget()
        chart_fl.addWidget(self.bar_chart)
        self.body_layout.addWidget(chart_frame)

        # ── تب‌ها ──
        tabs = QTabWidget()
        tabs.setStyleSheet(f"""
            QTabWidget::pane {{
                background: white;
                border: 1px solid {ModernStyle.BORDER};
                border-radius: 14px;
                top: -1px;
            }}
            QTabBar::tab {{
                background: #F0EEFF;
                color: {ModernStyle.TEXT_SECONDARY};
                padding: 10px 22px;
                margin: 2px 3px;
                border-radius: 10px 10px 0 0;
                font-size: 12px;
                font-family: Tahoma;
                border: 1px solid {ModernStyle.BORDER};
                border-bottom: none;
                min-height: 36px;
            }}
            QTabBar::tab:selected {{
                background: white;
                color: {ModernStyle.PRIMARY};
                font-weight: bold;
                border-bottom: 2px solid {ModernStyle.PRIMARY};
            }}
            QTabBar::tab:hover:!selected {{ background: #E8E0FF; }}
        """)
        tabs.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        tabs.setMinimumHeight(380)

        tabs.addTab(self._build_monthly_tab(),   "📅  فروش ماهانه")
        tabs.addTab(self._build_products_tab(),  "🏆  محصولات برتر")
        tabs.addTab(self._build_customers_tab(), "👑  مشتریان برتر")
        tabs.addTab(self._build_inventory_tab(), "📦  گزارش انبار")

        self.body_layout.addWidget(tabs, stretch=1)

    # ── helpers ──────────────────────────────
    @staticmethod
    def _lbl(text, color=None, size=12):
        l = QLabel(text)
        l.setFont(QFont("Tahoma", size))
        style = "background:transparent;border:none;"
        if color:
            style += f"color:{color};"
        l.setStyleSheet(style)
        return l

    @staticmethod
    def _make_table(cols, headers, stretch_col=None):
        t = QTableWidget()
        t.setColumnCount(cols)
        t.setHorizontalHeaderLabels(headers)
        t.verticalHeader().setVisible(False)
        t.setAlternatingRowColors(True)
        t.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        t.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        t.verticalHeader().setDefaultSectionSize(42)
        t.horizontalHeader().setMinimumSectionSize(80)
        t.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        t.setShowGrid(True)
        t.setStyleSheet(f"""
            QTableWidget {{
                border: none;
                background: white;
                gridline-color: {ModernStyle.BORDER_LIGHT};
                font-size: 12px;
                font-family: Tahoma;
            }}
            QTableWidget::item {{
                padding: 0 10px;
                min-height: 42px;
                color: {ModernStyle.TEXT_PRIMARY};
                border-bottom: 1px solid {ModernStyle.BORDER_LIGHT};
            }}
            QTableWidget::item:selected {{
                background: #EDE7FF;
                color: {ModernStyle.PRIMARY};
            }}
            QTableWidget::item:alternate {{
                background: #F8F9FC;
            }}
            QHeaderView::section {{
                background: {ModernStyle.PRIMARY};
                color: white;
                padding: 10px 10px;
                border: none;
                font-weight: bold;
                font-size: 12px;
                font-family: Tahoma;
                min-height: 42px;
            }}
        """)
        hdr = t.horizontalHeader()
        for i in range(cols):
            hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        if stretch_col is not None:
            hdr.setSectionResizeMode(stretch_col, QHeaderView.ResizeMode.Stretch)
        else:
            hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        return t

    @staticmethod
    def _export_btn(label, slot):
        btn = QPushButton(label)
        btn.setFixedHeight(36)
        btn.setMaximumWidth(160)
        btn.setStyleSheet(f"""
            QPushButton {{
                background: {ModernStyle.SUCCESS};
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 12px;
                font-family: Tahoma;
                padding: 0 14px;
            }}
            QPushButton:hover {{ background: #00A381; }}
        """)
        btn.clicked.connect(slot)
        return btn

    # ── ساخت تب‌ها ───────────────────────────
    def _tab_container(self):
        w = QWidget()
        w.setStyleSheet("background:white;")
        lay = QVBoxLayout(w)
        lay.setContentsMargins(16, 14, 16, 14)
        lay.setSpacing(12)
        return w, lay

    def _build_monthly_tab(self):
        w, lay = self._tab_container()
        lay.addWidget(self._export_btn("📥 خروجی Excel", self._export_monthly),
                      alignment=Qt.AlignmentFlag.AlignLeft)
        self.monthly_table = self._make_table(
            4,
            ["ماه", "تعداد سفارش", "مبلغ کل (تومان)", "میانگین سفارش (تومان)"],
            stretch_col=0
        )
        lay.addWidget(self.monthly_table, stretch=1)
        return w

    def _build_products_tab(self):
        w, lay = self._tab_container()
        lay.addWidget(self._export_btn("📥 خروجی Excel", self._export_products),
                      alignment=Qt.AlignmentFlag.AlignLeft)
        self.products_table = self._make_table(
            4,
            ["رتبه", "نام محصول", "تعداد فروخته‌شده", "درآمد کل (تومان)"],
            stretch_col=1
        )
        lay.addWidget(self.products_table, stretch=1)
        return w

    def _build_customers_tab(self):
        w, lay = self._tab_container()
        lay.addWidget(self._export_btn("📥 خروجی Excel", self._export_customers),
                      alignment=Qt.AlignmentFlag.AlignLeft)
        self.customers_table = self._make_table(
            5,
            ["رتبه", "نام مشتری", "موبایل", "تعداد سفارش", "کل خرید (تومان)"],
            stretch_col=1
        )
        lay.addWidget(self.customers_table, stretch=1)
        return w

    def _build_inventory_tab(self):
        w, lay = self._tab_container()
        toolbar = QHBoxLayout()
        toolbar.addWidget(self._export_btn("📥 خروجی Excel", self._export_inventory))
        toolbar.addStretch()
        toolbar.addWidget(self._lbl("فیلتر:", size=11))
        self.inv_filter = QComboBox()
        self.inv_filter.addItems(["همه", "کم‌موجودی", "بدون موجودی"])
        self.inv_filter.setFixedHeight(36)
        self.inv_filter.setFixedWidth(130)
        self.inv_filter.currentTextChanged.connect(self._filter_inventory)
        toolbar.addWidget(self.inv_filter)
        lay.addLayout(toolbar)

        self.inventory_table = self._make_table(
            7,
            ["کد", "نام کالا", "دسته‌بندی", "رنگ", "سایز", "موجودی", "وضعیت"],
            stretch_col=1
        )
        lay.addWidget(self.inventory_table, stretch=1)
        return w

    # ── بارگذاری داده ──────────────────────
    def load_all(self):
        year = self.year_combo.currentData()
        self._load_monthly(year)
        self._load_products()
        self._load_customers()
        self._load_inventory()

    def _load_monthly(self, year):
        try:
            data = self.db.get_monthly_sales(year)

            # به‌روزرسانی نمودار
            self.bar_chart.set_data(data)

            # محاسبه خلاصه
            total_rev   = sum(d['amount'] for d in data)
            total_cnt   = sum(d['count']  for d in data)
            avg_order   = int(total_rev / total_cnt) if total_cnt else 0
            top_month   = max(data, key=lambda d: d['amount'], default={'month': '—'})['month']

            self.card_total_orders.set_value(f"{total_cnt:,} سفارش")
            self.card_total_revenue.set_value(f"{int(total_rev):,} ت")
            self.card_avg_order.set_value(f"{avg_order:,} ت")
            self.card_top_month.set_value(top_month)

            # جدول
            self.monthly_table.setRowCount(0)
            bold = QFont("Tahoma", 11, QFont.Weight.Bold)

            for row_idx, d in enumerate(data):
                self.monthly_table.insertRow(row_idx)
                self._cell(self.monthly_table, row_idx, 0, d['month'])
                self._cell(self.monthly_table, row_idx, 1, f"{d['count']:,}", center=True)
                self._cell(self.monthly_table, row_idx, 2, f"{int(d['amount']):,}", center=True)
                avg = int(d['amount'] / d['count']) if d['count'] else 0
                self._cell(self.monthly_table, row_idx, 3, f"{avg:,}", center=True)

            # ردیف جمع
            r = self.monthly_table.rowCount()
            self.monthly_table.insertRow(r)
            for col, val in enumerate(["🔢 جمع کل", f"{total_cnt:,}", f"{int(total_rev):,}", ""]):
                item = QTableWidgetItem(val)
                item.setFont(bold)
                item.setBackground(QBrush(QColor("#EDE7FF")))
                if col > 0:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                self.monthly_table.setItem(r, col, item)

        except Exception as e:
            print(f"load_monthly error: {e}")

    def _load_products(self):
        try:
            data = self.db.get_top_products(20)
            self.products_table.setRowCount(0)
            medals = {0: "🥇", 1: "🥈", 2: "🥉"}
            for i, (name, qty, rev) in enumerate(data):
                r = self.products_table.rowCount()
                self.products_table.insertRow(r)
                rank = medals.get(i, str(i + 1))
                self._cell(self.products_table, r, 0, rank, center=True)
                self._cell(self.products_table, r, 1, name or '')
                self._cell(self.products_table, r, 2, f"{int(qty or 0):,}", center=True)
                self._cell(self.products_table, r, 3, f"{int(rev or 0):,}", center=True)
                if i == 0:
                    self._highlight_row(self.products_table, r, "#FFF8E1")
        except Exception as e:
            print(f"load_products error: {e}")

    def _load_customers(self):
        try:
            data = self.db.get_top_customers(20)
            self.customers_table.setRowCount(0)
            medals = {0: "🥇", 1: "🥈", 2: "🥉"}
            for i, (name, mobile, cnt, total) in enumerate(data):
                r = self.customers_table.rowCount()
                self.customers_table.insertRow(r)
                rank = medals.get(i, str(i + 1))
                self._cell(self.customers_table, r, 0, rank, center=True)
                self._cell(self.customers_table, r, 1, name or '')
                self._cell(self.customers_table, r, 2, mobile or '', center=True)
                self._cell(self.customers_table, r, 3, f"{int(cnt or 0):,}", center=True)
                self._cell(self.customers_table, r, 4, f"{int(total or 0):,}", center=True)
                if i == 0:
                    self._highlight_row(self.customers_table, r, "#FFF8E1")
        except Exception as e:
            print(f"load_customers error: {e}")

    def _load_inventory(self):
        try:
            self._inv_data = self.db.get_inventory_report()
            self._render_inventory(self._inv_data)
        except Exception as e:
            print(f"load_inventory error: {e}")

    def _filter_inventory(self, f):
        if not hasattr(self, '_inv_data'):
            return
        if f == "کم‌موجودی":
            data = [r for r in self._inv_data if r[8] == 'کم' and r[5] > 0]
        elif f == "بدون موجودی":
            data = [r for r in self._inv_data if r[5] == 0]
        else:
            data = self._inv_data
        self._render_inventory(data)

    def _render_inventory(self, data):
        self.inventory_table.setRowCount(0)
        for (code, name, cat, color, size, stock, price, min_alert, status) in data:
            r = self.inventory_table.rowCount()
            self.inventory_table.insertRow(r)
            self._cell(self.inventory_table, r, 0, code or '', center=True)
            self._cell(self.inventory_table, r, 1, name or '')
            self._cell(self.inventory_table, r, 2, cat or '', center=True)
            self._cell(self.inventory_table, r, 3, color or '', center=True)
            self._cell(self.inventory_table, r, 4, size or '', center=True)

            stock_item = QTableWidgetItem(str(stock or 0))
            stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            if stock == 0:
                stock_item.setBackground(QBrush(QColor("#FFEBEB")))
                stock_item.setForeground(QBrush(QColor(ModernStyle.DANGER)))
            elif status == 'کم':
                stock_item.setBackground(QBrush(QColor("#FFF8E1")))
                stock_item.setForeground(QBrush(QColor(ModernStyle.WARNING)))
            self.inventory_table.setItem(r, 5, stock_item)

            badge = "🔴 کم‌موجودی" if status == 'کم' else "✅ کافی"
            status_item = QTableWidgetItem(badge)
            status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            self.inventory_table.setItem(r, 6, status_item)

    # ── helpers جدول ─────────────────────────
    @staticmethod
    def _cell(table, row, col, text, center=False):
        item = QTableWidgetItem(str(text))
        if center:
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
        else:
            item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        table.setItem(row, col, item)

    @staticmethod
    def _highlight_row(table, row, color):
        for c in range(table.columnCount()):
            item = table.item(row, c)
            if item:
                item.setBackground(QBrush(QColor(color)))

    # ── خروجی Excel ──────────────────────────
    def _export_monthly(self):
        self._to_excel(self.monthly_table, "فروش_ماهانه")

    def _export_products(self):
        self._to_excel(self.products_table, "محصولات_برتر")

    def _export_customers(self):
        self._to_excel(self.customers_table, "مشتریان_برتر")

    def _export_inventory(self):
        self._to_excel(self.inventory_table, "گزارش_انبار")

    def _to_excel(self, table, prefix):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.warning(self, "خطا", "openpyxl نصب نیست:\npip install openpyxl")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "ذخیره Excel",
            f"{prefix}_{get_current_shamsi_date().replace('/', '-')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = prefix
            hfill  = PatternFill("solid", fgColor="6C5CE7")
            hfont  = Font(bold=True, color="FFFFFF", name="Tahoma")
            center = Alignment(horizontal="center", vertical="center")

            for col in range(table.columnCount()):
                h = table.horizontalHeaderItem(col)
                cell = ws.cell(1, col + 1, h.text() if h else f"col{col+1}")
                cell.fill = hfill; cell.font = hfont; cell.alignment = center
                ws.column_dimensions[cell.column_letter].width = 22

            for row in range(table.rowCount()):
                for col in range(table.columnCount()):
                    it = table.item(row, col)
                    ws.cell(row + 2, col + 1, it.text() if it else "").alignment = center

            wb.save(path)
            QMessageBox.information(self, "✅ موفق", f"فایل ذخیره شد:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "خطا", str(e))
